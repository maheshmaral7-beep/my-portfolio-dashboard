"""
import_holdings.py — converts manually-exported holdings/mutual-fund statements
(the .xlsx files you download from Zerodha/Groww/Console) into the JSON files
the dashboard reads: portfolio.json and mutual_funds.json.

This is the "manual import" path for Zerodha/Groww holdings and mutual funds,
used until (or instead of) a live API connection.

USAGE:
  python3 import_holdings.py --stocks Stocks_Holdings_Statement_....xlsx --mf Mutual_Funds_....xlsx

Either argument can be omitted if you only want to update one of the two.

PRIVACY NOTE: this script deliberately does NOT read or output your name, PAN,
mobile number, or folio numbers into the generated JSON — only portfolio figures.
"""

import argparse
import json
import re
from datetime import datetime, timezone

import openpyxl


def parse_stocks(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    summary = {}
    holdings = []
    as_of = None
    header_seen = False

    for row in rows:
        if row[0] and isinstance(row[0], str) and "Holdings statement for stocks as on" in row[0]:
            m = re.search(r"as on ([\d-]+)", row[0])
            if m:
                as_of = m.group(1)
        if row[0] == "Invested Value":
            summary["invested_value"] = float(row[1])
        elif row[0] == "Closing Value":
            summary["closing_value"] = float(row[1])
        elif row[0] == "Unrealised P&L":
            summary["total_pl"] = float(row[1])
        elif row[0] == "Stock Name":
            header_seen = True
            continue
        elif header_seen and row[0]:
            name, isin, qty, avg_buy, buy_val, close_price, close_val, pl = row[:8]
            no_price_data = (close_price in (0, None)) and (avg_buy in (0, None))
            holdings.append({
                "name": name,
                "isin": isin,
                "quantity": qty,
                "avg_buy_price": avg_buy,
                "buy_value": buy_val,
                "closing_price": close_price,
                "closing_value": close_val,
                "pl": pl,
                "no_price_data": no_price_data,
            })

    priced_holdings = [h for h in holdings if not h["no_price_data"]]
    top_contributor = max(priced_holdings, key=lambda h: h["pl"], default=None)
    top_detractor = min(priced_holdings, key=lambda h: h["pl"], default=None)

    invested = summary.get("invested_value", 0)
    total_pl = summary.get("total_pl", 0)
    total_pl_pct = round((total_pl / invested) * 100, 2) if invested else None

    return {
        "status": "connected",
        "source": "manual_export",
        "as_of": as_of,
        "imported_at": datetime.now(timezone.utc).isoformat(),
        "portfolio_value": summary.get("closing_value"),
        "invested_value": invested,
        "total_pl": total_pl,
        "total_pl_pct": total_pl_pct,
        "day_pl": None,
        "day_pl_note": "Not available from a static export -- would need holdings cross-referenced with live prices from poll.py.",
        "top_contributor": {"name": top_contributor["name"], "pl": top_contributor["pl"]} if top_contributor else None,
        "top_detractor": {"name": top_detractor["name"], "pl": top_detractor["pl"]} if top_detractor else None,
        "holdings": holdings,
    }


def parse_mutual_funds(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    summary = {}
    funds = []
    as_of = None
    header_seen = False

    for i, row in enumerate(rows):
        if row[0] == "HOLDING SUMMARY":
            header_row = rows[i + 3]
            value_row = rows[i + 4]
            summary["total_investments"] = float(value_row[0])
            summary["current_value"] = float(value_row[1])
            summary["profit_loss"] = float(value_row[2])
            summary["profit_loss_pct"] = float(str(value_row[3]).replace("%", ""))
            summary["xirr_pct"] = float(str(value_row[4]).replace("%", ""))
        if row[0] and isinstance(row[0], str) and row[0].startswith("HOLDINGS AS ON"):
            m = re.search(r"HOLDINGS AS ON ([\d-]+)", row[0])
            if m:
                as_of = m.group(1)
        if row[0] == "Scheme Name":
            header_seen = True
            continue
        elif header_seen and row[0]:
            scheme, amc, category, sub_category, folio, source, units, invested, current, returns, xirr = row[:11]
            funds.append({
                "scheme_name": scheme,
                "amc": amc,
                "category": category,
                "sub_category": sub_category,
                "source": source,
                "units": units,
                "invested_value": float(invested) if invested is not None else None,
                "current_value": float(current) if current is not None else None,
                "returns": returns,
                "xirr_pct": float(str(xirr).replace("%", "")) if xirr is not None else None,
            })

    allocation_value = {}
    for f in funds:
        key = f["sub_category"] or "Other"
        allocation_value[key] = allocation_value.get(key, 0) + (f["current_value"] or 0)
    total_current = sum(allocation_value.values()) or 1
    allocation_pct = {k: round(v / total_current * 100, 1) for k, v in allocation_value.items()}

    return {
        "status": "connected",
        "source": "manual_export",
        "as_of": as_of,
        "imported_at": datetime.now(timezone.utc).isoformat(),
        "total_invested": summary.get("total_investments"),
        "current_value": summary.get("current_value"),
        "gain_loss": summary.get("profit_loss"),
        "gain_loss_pct": summary.get("profit_loss_pct"),
        "xirr_pct": summary.get("xirr_pct"),
        "allocation_pct": allocation_pct,
        "funds": funds,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--stocks", help="Path to the stocks holdings statement .xlsx")
    parser.add_argument("--mf", help="Path to the mutual funds statement .xlsx")
    args = parser.parse_args()

    if args.stocks:
        result = parse_stocks(args.stocks)
        with open("portfolio.json", "w") as f:
            json.dump(result, f, indent=2, default=str)
        print("Saved portfolio.json")

    if args.mf:
        result = parse_mutual_funds(args.mf)
        with open("mutual_funds.json", "w") as f:
            json.dump(result, f, indent=2, default=str)
        print("Saved mutual_funds.json")


if __name__ == "__main__":
    main()
